# pysimpleGUI for conversion of multiple Med Associate text files into one Excel file

# Import required libraries
import PySimpleGUI as sg
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from tempfile import TemporaryFile
from pathlib import Path
import re

# validate that the file paths are entered correctly
def is_valid_path(filepath):
    if filepath and Path(filepath).exists():
        return True
    sg.popup_error("A selected file path is incorrect or has been left empty.")
    return False

# window appears when the program successfully completes
def nom_window():
    layout = [[sg.Text("\n"
    " All Systems Nominal  \n"
    "\n"
    "")]]
    window = sg.Window((""), layout, modal=True)
    choice = None
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
    window.close()
    
# Define the location of the directory
def extract_values_from_excel(input_folder, output_folder):
    name = Path(output_folder)

    # Change the directory
    os.chdir(input_folder)
    print(input_folder)

    # Create a new Excel file to store the extracted values
    output_workbook = openpyxl.Workbook()
    sheet1 = output_workbook.active

    sheet1['A1'] = input_folder
    sheet1['A2'] = 'Subj_ID'
    sheet1['B2'] = 'Time_Sum'
    sheet1['C2'] = 'Bout_Sum'
    sheet1['D2'] = 'Av_Bout_Tm'

    row_var = 3 # Start writing data from row 3

    # creation of a maximum value for the progress bar function
    input_folder  = values["-IN-"]
    prog_bar_max_val = 1
    os.chdir(input_folder)
    for i in os.listdir():
        prog_bar_max_val += 1
    max = prog_bar_max_val
    prog_bar_update_val = 0

    for files in os.listdir(input_folder):
        filepath = os.path.join(input_folder, files)

        # Load the Excel file
        workbook = openpyxl.load_workbook(filepath)
        
        # Extract values from 'SumData' Sheet
        worksheet = workbook['SumData']
        
        bout_column = 'B'
        last_row = None

        # Iterate through the rows in the specified column from the bottom to the top
        for row in range(worksheet.max_row, 1, -1):
            cell = worksheet[f'{bout_column}{row}']
            # updating prog bar with each file compiled
            window["-Progress_BAR-"].update(max = max, current_count=int(prog_bar_update_val))
            if cell.value is not None:
                last_row = cell.value
                # Extract values from two specific cells
                tot_time = worksheet['A2'].value
                tot_bout = worksheet['B'+str(last_row+1)].value
                av_bt_tm = round((tot_time/tot_bout),1)
                prog_bar_update_val += 1
                break
            else: 
                tot_time = 0
            tot_bout = 0
            av_bt_tm = 0
            prog_bar_update_val += 1
            
        

        # iterate over all the files in the directory 
        #prog_bar_update_val = 0
        #for files in os.listdir():
            #prog_bar_update_val += 1

        sheet1['B'+str(row_var)] = tot_time
        sheet1['C'+str(row_var)] = tot_bout
        sheet1['D'+str(row_var)] = av_bt_tm
        print(tot_time)
        print(tot_bout)
        print(av_bt_tm)
        row_var += 1

    row = 3  # Start writing data from row 1
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            # Extract numbers from the file name using regular expression
            numbers = re.findall(r'\d+', filename)
            
            if numbers:
                # Write the extracted numbers to the Excel file
                for num in numbers:
                    sheet1.cell(row=row, column=1, value=int(num))
                    row += 1

    name = input_folder+'.xlsx'
    output_workbook.save(name)
    print('saved')
    # last prog bar addition indicating the end of the program run
    window["-Progress_BAR-"].update(current_count=int(prog_bar_update_val +1))

    # window telling the user the program functioned correctly
    nom_window()   

# creation of a maximum value for the progress bar function
def bar_max(input_folder):
    prog_bar_max_val = 0
    os.chdir(input_folder)
    for i in os.listdir():
        prog_bar_max_val += 1

# main GUI creation and GUI elements
sg.theme('DarkBlue2')

layout = [
    [sg.Text("Select the folder containing the\n"
             "behavioral Excel data files                         \n" 
             "to be compiled into one Excel file."),
    sg.Input(key="-IN-"),
    sg.FolderBrowse()],

    [sg.Text("Select a file to store the new Excel file.\n"
                "Data will be copied & transferred to this file.\n"),
    sg.Input(key="-OUT-"),
    sg.FolderBrowse()],

    [sg.Exit(), sg.Button("Press to compile bout data into new Excel file"), 
    sg.Text("eBot's progress..."),
    sg.ProgressBar(20, orientation='horizontal', size=(15,10), 
                border_width=4, bar_color=("Blue", "Grey"),
                key="-Progress_BAR-")]
    
]

# create the window
window = sg.Window("Welcome to eBot's behavioral Excel data compiler!", layout)

# create an event loop
while True:
    event, values = window.read()
    # end program if user closes window
    if event == "Exit" or event == sg.WIN_CLOSED:
        break
    if event == "Press to compile bout data into new Excel file":
        # check file selections are valid
        if (is_valid_path(values["-IN-"])) and (is_valid_path(values["-OUT-"])):

            extract_values_from_excel(
            input_folder  = values["-IN-"],
            output_folder = values["-OUT-"])   

window.close