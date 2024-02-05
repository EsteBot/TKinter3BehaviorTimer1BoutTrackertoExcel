import threading
import time
import keyboard
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Dictionary to store the time
timers = {}

t1_lst = []
t2_lst = []
t3_lst = []

bout_lst = []

t1 = 0
t2 = 0
t3 = 0

def start_timer(timer_name):
    if timer_name in timers:
        timer = timers[timer_name]
        if 'start_time' not in timer:
            start_time = time.time()
            timer['start_time'] = start_time
            print(f"Timer '{timer_name}' restarted.")
        else:
            print(f"Timer '{timer_name}' is already running.")
    else:
        start_time = time.time()
        timers[timer_name] = {
            'start_time': start_time,
            'total_time': 0
        }
        print(f"Timer '{timer_name}' started.")

def stop_timer(timer_name):
    global t1, t2, t3  # Declare the variables as global
    if timer_name in timers:
        timer = timers[timer_name]
        if 'start_time' in timer:
            elapsed_time = time.time() - timer['start_time']
            timer['total_time'] += elapsed_time
            del timer['start_time']
            print(f"Timer '{timer_name}' stopped. Bout time: {round(elapsed_time, 1)} seconds.")
            if timer_name == "Timer 1":
                t1 += 1
                t1_lst.append(round(elapsed_time, 1))
            elif timer_name == "Timer 2":
                t2 += 1
                t2_lst.append(round(elapsed_time, 1))
            elif timer_name == "Timer 3":
                t3 += 1
                t3_lst.append(round(elapsed_time, 1))

        else:
            print(f"Timer '{timer_name}' is already stopped.")
    else:
        print(f"No timer found with the name '{timer_name}'.")


def assign_timer_keys(timer_name, start_key, stop_key):
    keyboard.on_press_key(start_key, lambda _: start_timer(timer_name))
    keyboard.on_press_key(stop_key, lambda _: stop_timer(timer_name))
    print(f"Assigned keys '{start_key}' and '{stop_key}' to timer '{timer_name}'.")

def assign_keys_timers():
    timer_names = ["Timer 1", "Timer 2", "Timer 3"]
    start_keys = ["q", "a", "o"]
    stop_keys = ["w", "s", "p"]

    for i in range(len(timer_names)):
        assign_timer_keys(timer_names[i], start_keys[i], stop_keys[i])

def display_total_time():
    print("Press an assigned 'start' key to begin or 'esc' to exit and 't' to save data.")
    for timer_name, timer_data in timers.items():
        total_time = timer_data['total_time']
        print(f"Timer '{timer_name}': {total_time} seconds")


filename = input("Enter a file name for the generated Excel data file:\n")
assign_keys_timers()

# Start a separate thread to display the total elapsed time
timer_thread = threading.Thread(target=display_total_time)
timer_thread.start()

def excel_func():
    wb = Workbook()
    ws = wb.active

    # Create a new worksheet named 'SumData'
    sheet1 = wb.create_sheet('SumData')
    sheet1 = wb['SumData']

    sheet1['A1'] = 'Timer 1'
    sheet1['A2'] = sum(t1_lst)
    sheet1['B1'] = 'bout'
    sheet1['C1'] = 'sec'
    for i, value in enumerate(t1_lst, start=2):
        sheet1.cell(row=i, column=3, value=value)
        sheet1.cell(row=i, column=2, value=i-1)

    sheet1['E1'] = 'Timer 2'
    sheet1['E2'] = sum(t2_lst)
    sheet1['F1'] = 'bout'
    sheet1['G1'] = 'sec'
    for i, value in enumerate(t2_lst, start=2):
        sheet1.cell(row=i, column=7, value=value)
        sheet1.cell(row=i, column=6, value=i-1)  

    sheet1['I1'] = 'Timer 3'
    sheet1['I2'] = sum(t3_lst)
    sheet1['J1'] = 'bout'
    sheet1['K1'] = 'sec'
    for i, value in enumerate(t3_lst, start=2):
        sheet1.cell(row=i, column=11, value=value)
        sheet1.cell(row=i, column=10, value=i-1)


    # Save the workbook
    wb.save(filename+'.xlsx')

    print("Bout Timer Data saved to "+filename+".xlsx\n")

def summary_func():
    # Calculate and print the sum of elapsed times for each timer
    print("Sum of bout times for each timer:")
    for timer_name, timer_data in timers.items():
        total_elapsed_time = timer_data['total_time']
        print(f"Total '{timer_name}' time: {round(total_elapsed_time, 1)} seconds")
        if timer_name == "Timer 1":
            print(f'Total bouts: {t1}')
        elif timer_name == "Timer 2":
            print(f'Total bouts: {t2}')
        else:
            print(f'Total bouts: {t3}')

keyboard.wait('t')
summary_func()
excel_func()

# Keep the program running
keyboard.wait('esc')
